import re
from openpyxl import load_workbook
from Database import Database
from CrawlerHelpScripts import logger
import time

log = logger.createLogger("websiteParserLogger", "websiteParser")

class Parser():

    def main(self):
        log.info(str(time.asctime( time.localtime(time.time()) ))+" starting websiteParser")
        #Import the excel file with websites
        wb=load_workbook(r'C:\test.xlsx', use_iterators = True)
        ws=wb.get_sheet_by_name('Sheet1')
        regex = '^[a-zA-Z0-9\-\.]+\.[a-zA-Z]{2,3}(/\S*)?$' #This is the regex to check for websites
        
        self.db = Database.Queries()
        self.db.openConnection()
        
        #Iterate trough all rows
        for row in ws.iter_rows(row_offset=1):
            for cell in row:
                #If the column == A, check if there's a website value
                if cell.column == 'A':
                    try:
                        self.match = re.match(regex, cell.internal_value)
                        if self.match:
                            self.match = 'OK'
                    except:
                        pass
                #If the column is a website value, crawl data from file
                if self.match == 'OK':
                    self.getInfo(cell)
                    
        #Close connection after all is done            
        self.db.closeConnection()
        log.info(str(time.asctime( time.localtime(time.time()) ))+" Website parsing done")
                    
    def getInfo(self,cell):              
        if cell.column == 'A':
            self.weburl = cell.internal_value
            
            #Extract the name from the website
            begin = self.weburl.find('.')+1
            end = self.weburl[begin:].find('.')+begin
            #If the end variable equals begin-1, there is no www. in it. So then just extract till the 1st '.'
            if end == begin-1:
                self.name = self.weburl[:begin].title()
            else:
                self.name = self.weburl[begin:end].title()
              
        elif cell.column == 'B':
            self.country = cell.internal_value
        
        elif cell.column == 'C':
            self.sender = str(cell.internal_value)
            #If the string contains a ',' it means that it contains multiple values, strip these
            if self.sender == None or self.sender == '-':
                self.sender = None
            elif self.sender.find(',') != -1:
                self.sender = self.split(self.sender)
                
        elif cell.column == 'D':
            self.shipping_price = str(cell.internal_value)
            
            try:
                #Replace , with a '.' to avoid truncated data
                self.shipping_price = self.shipping_price.replace(',', '.')
        
                decimalFinder = self.shipping_price.find('.') 
                if decimalFinder == -1: #If it is a round number, add decimals so the db won't be updated unnecessary
                    self.shipping_price = self.shipping_price + '.00'
            except:
                pass
            
        elif cell.column == 'E':
            self.mark = str(cell.internal_value)
            #If the string contains a ',' it means that it contains multiple values, strip these
            if self.mark == None or self.mark == '-':
                self.mark = None
            elif self.mark.find(',') != -1:
                self.mark = self.split(self.mark)
            
        if cell.column == 'G':
            self.payment_methods = str(cell.internal_value)
            #If the string contains a ',' it means that it contains multiple values, strip these
            if self.payment_methods == None or self.payment_methods == '-':
                self.payment_methods = None
            elif self.payment_methods.find(',') != -1:
                self.payment_methods = self.split(self.payment_methods)
            
            self.match = None #Set this to None to avoid none-url rows from being parsed
            
            #Save the data to the database
            self.db.saveWebsiteInfo(self.weburl, self.name, self.country, self.shipping_price, self.sender, self.mark, self.payment_methods)
    
    def split(self, source):#Used to put multiple values from eg. payment_methods in a list
        string = source.replace(" ", "")
        li = string.split(",")
        
        return li
        
                    
p = Parser()
p.main()             
    
