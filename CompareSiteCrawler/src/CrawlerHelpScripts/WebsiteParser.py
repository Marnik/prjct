import re
from openpyxl import load_workbook
from Database import Database

#Import the excel file with websites
wb=load_workbook(r'C:\test.xlsx', use_iterators = True)
ws=wb.get_sheet_by_name('Sheet1')
regex = '^[a-zA-Z0-9\-\.]+\.[a-zA-Z]{2,3}(/\S*)?$' #This is the regex to check for websites

db = Database.Queries()
db.openConnection()

#Iterate trough all rows
for row in ws.iter_rows(row_offset=1):
    for cell in row:
        #If the column == A, check if there's a website value
        if cell.column == 'A':
            try:
                match = re.match(regex, cell.internal_value)
                if match:
                    match = 'OK'
            except:
                pass
        #If the column is a website value, crawl data from file
        if match == 'OK':
            if cell.column == 'A':
                weburl = cell.internal_value
                #Extract the name from the website
                begin = weburl.find('.')+1
                end = weburl[begin:].find('.')+begin
                
                #If the end variable equals begin-1, there is no www. in it. So then just extract till the 1st '.'
                if end == begin-1:
                    name = weburl[:begin].title()
                else:
                    name = weburl[begin:end].title()
            elif cell.column == 'B':
                country = cell.internal_value
            elif cell.column == 'C':
                sender = cell.internal_value
            elif cell.column == 'D':
                shipping_price = cell.internal_value
                
                try:
                    #Replace , with a '.' to avoid truncated data
                    shipping_price = shipping_price.replace(',', '.')
        
                    decimalFinder = shipping_price.find('.') 
                    if decimalFinder == -1: #If it is a round number, add decimals so the db won't be updated unnecessary
                        shipping_price = shipping_price + '.00'
                except:
                    pass
            elif cell.column == 'E':
                mark = cell.internal_value
            if cell.column == 'G':
                payment_methods = cell.internal_value
                match = None
                
                #Save the data to the website
                db.saveWebsites(weburl, name, country, shipping_price)
                
db.closeConnection()
                
                

